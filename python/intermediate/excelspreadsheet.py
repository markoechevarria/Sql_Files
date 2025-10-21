import openpyxl

wb = openpyxl.load_workbook('example.xlsx')
print(type(wb))

print(wb.sheetnames)
sheet1 = wb['Sheet1']
print(type(sheet1))
print(sheet1.title)

print( sheet1['A1'] )
print( sheet1['B1'].value  )

cell = sheet1['A1']

print( 'Row %s, Column %s is %s' % (cell.row, cell.column, cell.value) )
print( 'Cell %s is %s' % (cell.coordinate, cell.value) )
print( sheet1['C1'].value )

# .cell( row, column )

print(sheet1.cell(row=1, column=2).value)
for i in range(1,6):
	print(i, sheet1.cell(row=i, column=2).value )

print('# Max_row: ', sheet1.max_row)
print('# Max_column: ', sheet1.max_column)

from openpyxl.utils import get_column_letter, column_index_from_string
print( get_column_letter(1) )
print( get_column_letter(27)  )
print( get_column_letter(sheet1.max_column) )
print( column_index_from_string('AA') )

# Getting rows and columns from the sheets

print( sheet1['A1':'C3'] )
tupleSheet = sheet1['A1':'C3']
for rowOfCellObjects in tupleSheet:
	for cellObj in rowOfCellObjects:
		print( cellObj.coordinate, cellObj.value )
	print('--- END OF ROW ---')

listSheet = list(sheet1.columns)[1]
print(listSheet)
for cellObj in listSheet:
	print(cellObj.value)

import openpyxl, pprint

print('Opening workbook...')
wb = openpyxl.load_workbook('censuspopdata.xlsx')
sheet = wb['Population by Census Tract']
print(sheet['A1'])

countyData = {}

for row in range(2,sheet.max_row+1):
	state = sheet['B' + str(row)].value
	county = sheet['C' + str(row)].value
	pop = sheet['D' + str(row)].value

	countyData.setdefault(state, {})
	countyData[state].setdefault(county, {'tracts': 0, 'pop': 0})
	countyData[state][county]['tracts'] += 1
	countyData[state][county]['pop'] += int(pop)

resultFile = open('census2010.py', 'w')
resultFile.write('allData = ' + pprint.pformat(countyData) )
resultFile.close()
print('Done.')


# Creating and Saving Excel Documents

wb = openpyxl.Workbook()
wb.sheetnames

sheet = wb.active
print(sheet.title)
sheet.title = 'Spam Bacon Eggs Sheet'
print(wb.sheetnames)
wb.save('example_copy.xlsx')

# Creating and Removing Sheets

wb = openpyxl.Workbook()
wb.sheetnames

print(wb.sheetnames)

wb.create_sheet()

print(wb.sheetnames)

wb.create_sheet(index=0, title = 'First Sheet')

print(wb.sheetnames)

wb.create_sheet(index=2, title= 'Middle Sheet')

print(wb.sheetnames)

del wb['Middle Sheet']
del wb['Sheet1']

print(wb.sheetnames)


# Wriging values to cells

wb = openpyxl.Workbook()
sheet = wb['Sheet']
sheet['A1'] = 'Hello, world'
print(sheet['A1'].value)

# updateProduce.py

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']
PRICE_UPDATES = {'Garlic': 3.07,
		 'Celery': 1.19,
		 'Lemon': 1.27 }

for rowNum in range(2, sheet.max_row):
	produceName = sheet.cell(row=rowNum, column=1).value
	if produceName in PRICE_UPDATES:
		sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]

wb.save('updateProduceSales.xlsx')

# Setting the font style of cells

from openpyxl.styles import Font
wb = openpyxl.Workbook()
sheet = wb['Sheet']
italic24Font = Font(size = 24, italic = True)
sheet['A1'].font = italic24Font
sheet['A1'] = 'Hello world!'
wb.save('stylex.xlsx')
# Font(name='', bold=True, italic=True, size=25)

"""
	sheet['A1'] = 200
	sheet['A2'] = 300
	sheet['A3'] = '=SUM(A1:A2)'
"""

# Setting row height and column width

"""
sheet.row_dimensions[1].height = 70
sheet.column_dimensions['B'].width = 20
"""

# Merging and unmerging cells

"""
	sheet.merge_cells('A1:D3')
	sheet.unmerge_cells('A1:D3')
"""

# Freezing panes

"""
	sheet.freeze_panes = 'A2'
"""

# Creating a Chart

"""
>>> import openpyxl
>>> wb = openpyxl.Workbook()
>>> sheet = wb.active
>>> for i in range(1, 11): # create some data in column A
...     sheet['A' + str(i)] = i
...
>>> refObj = openpyxl.chart.Reference(sheet, min_col=1, min_row=1, max_col=1,
max_row=10)
>>> seriesObj = openpyxl.chart.Series(refObj, title='First series')

>>> chartObj = openpyxl.chart.BarChart()
>>> chartObj.title = 'My Chart'
>>> chartObj.append(seriesObj)

>>> sheet.add_chart(chartObj, 'C5')
>>> wb.save('sampleChart.xlsx')
"""
